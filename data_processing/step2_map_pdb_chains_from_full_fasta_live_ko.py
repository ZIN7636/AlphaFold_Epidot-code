r"""
step2_map_pdb_chains_from_full_fasta_live_ko.py
===============================================

핵심 옵션
---------
--verbose                 : 실시간 로그 출력
--progress-every N        : N개 체인마다 진행 메시지
--log-file path           : 로그를 파일로도 저장(콘솔+파일 동시 기록)
--flush-every N           : CSV를 N개 체인마다 flush(기본 1 → 매 체인마다 flush)
--append                  : 기존 CSV에 이어쓰기(중단 후 재시작 지원)

사용 예
-------
python step2_map_pdb_chains_from_full_fasta_live_ko.py ^
  --per-parent "d:\final\aws\per_parent_epitopes_clean_uniprot_label.xlsx" ^
  --pdb-fasta  "d:\final\aws\pdb_chains.fasta" ^
  --out-best   "d:\final\aws\mapping_best.csv" ^
  --out-all    "d:\final\aws\mapping_all.csv" ^
  --out-unmapped "d:\final\aws\unmapped_chains.csv" ^
  --verbose --progress-every 1 --log-file "d:\final\aws\map_run.log" --flush-every 1
-------------------------
"""

import re
import csv
import sys
import argparse
from pathlib import Path
from collections import defaultdict, Counter

# ------------------------- 로깅 -------------------------
class TeeLogger:
    def __init__(self, log_path=None):
        self.log_fp = None
        if log_path:
            p = Path(log_path); p.parent.mkdir(parents=True, exist_ok=True)
            self.log_fp = open(p, "a", encoding="utf-8", newline="")
    def log(self, msg, *, flush=True):
        print(msg, file=sys.stdout)
        if self.log_fp:
            print(msg, file=self.log_fp)
        if flush:
            sys.stdout.flush()
            if self.log_fp: self.log_fp.flush()
    def close(self):
        if self.log_fp:
            self.log_fp.close()

# ------------------------- 유틸 -------------------------
def clean_seq(s: str) -> str:
    """아미노산 서열을 대문자 알파벳만 남기도록 정리"""
    return re.sub(r"[^A-Za-z]", "", (s or "")).upper()

def read_csv_dicts(path: Path):
    """CSV를 dict 리스트로 읽기"""
    last=None
    for enc in ("utf-8-sig","utf-8","cp949","latin-1"):
        try:
            with open(path, newline="", encoding=enc, errors="ignore") as f:
                rdr = csv.DictReader(f)
                return [r for r in rdr]
        except Exception as e:
            last=e
    raise RuntimeError(f"CSV 읽기 실패: {path} ({last})")

def read_xlsx_dicts(path: Path, sheet_name=None):
    """openpyxl로 XLSX를 dict 리스트로 읽기"""
    try:
        import openpyxl 
    except Exception as e:
        raise RuntimeError("XLSX를 읽으려면 openpyxl이 필요합니다. "
                           "pip install openpyxl 또는 파일을 CSV로 저장해 주세요.") from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = {}
        for i, val in enumerate(r):
            key = header[i] if i < len(header) else f"col{i+1}"
            d[key] = val
        out.append(d)
    return out

def read_per_parent(path: Path):
    """per-parent 파일(CSV/XLSX) 읽고 (parent_name, full_fasta) 리스트 반환"""
    suffix = path.suffix.lower()
    if suffix==".csv":
        rows = read_csv_dicts(path)
    elif suffix in (".xlsx",".xlsm",".xltx",".xltm"):
        rows = read_xlsx_dicts(path)
    else:
        raise RuntimeError(f"지원하지 않는 확장자: {path.suffix} (csv/xlsx만 지원)")

    # 컬럼 동의어 매핑
    def norm(s): return re.sub(r"[^a-z0-9]+","_", (s or "").lower()).strip("_")
    parent_keys = {"parent_name","parent","molecule_parent","epitope_molecule_parent","epitope_-_molecule_parent"}
    fasta_keys  = {"full_fasta","fasta","sequence","full_sequence","uniprot_fasta"}

    if not rows:
        return []

    colmap = {norm(k): k for k in rows[0].keys()}
    c_parent = None
    for k in parent_keys:
        if k in colmap: c_parent = colmap[k]; break
    c_fasta = None
    for k in fasta_keys:
        if k in colmap: c_fasta = colmap[k]; break
    if not c_parent or not c_fasta:
        raise RuntimeError(f"필수 컬럼이 없습니다. (헤더: {list(rows[0].keys())})\n"
                           f"필요: parent_name 동의어 {sorted(parent_keys)}, full_fasta 동의어 {sorted(fasta_keys)}")

    out = []
    for r in rows:
        pname = str(r.get(c_parent) or "").strip()
        fseq  = clean_seq(r.get(c_fasta) or "")
        if pname and fseq:
            out.append((pname, fseq))
    return out

def read_fasta(path: Path):
    """간단 FASTA 파서: {header: sequence}"""
    seqs = {}
    header=None
    buf=[]
    with open(path, encoding="utf-8") as f:
        for line in f:
            line=line.strip()
            if not line: continue
            if line.startswith(">"):
                if header is not None:
                    seqs[header]="".join(buf)
                header=line[1:].strip()
                buf=[]
            else:
                buf.append(re.sub(r"\s+","",line))
        if header is not None:
            seqs[header]="".join(buf)
    for k in list(seqs.keys()):
        seqs[k]=clean_seq(seqs[k])
    return seqs

# ------------------------- k-mer 인덱싱/매핑 -------------------------
def build_kmer_index(parents, k=7, max_positions_per_kmer=2000):
    index = defaultdict(lambda: defaultdict(list))
    for pi, (_, seq) in enumerate(parents):
        L=len(seq)
        if L<k: continue
        for i in range(L-k+1):
            kmer = seq[i:i+k]
            if kmer.count(kmer[0]) == k:  
                continue
            d = index[kmer]
            lst = d[pi]
            if len(lst) < max_positions_per_kmer:
                lst.append(i)
    return index

def vote_offsets_for_chain(chain_seq, index, k):
    votes = defaultdict(Counter)
    kmer_hits = Counter()
    L = len(chain_seq)
    if L < k: 
        return votes, kmer_hits
    for i in range(L-k+1):
        kmer = chain_seq[i:i+k]
        d = index.get(kmer)
        if not d: 
            continue
        for pi, poslist in d.items():
            kmer_hits[pi] += len(poslist)
            for ppos in poslist:
                off = ppos - i
                votes[pi][off] += 1
    return votes, kmer_hits

def eval_alignment(parent_seq, chain_seq, offset):
    pL=len(parent_seq); cL=len(chain_seq)
    parent_start = offset if offset>0 else 0
    chain_start  = -offset if offset<0 else 0
    max_len = min(pL - parent_start, cL - chain_start)
    if max_len <= 0:
        return 0, 0, 0.0, 0.0, 0.0, 0, 0, 0
    matches = 0
    for i in range(max_len):
        if parent_seq[parent_start+i] == chain_seq[chain_start+i]:
            matches += 1
    identity = matches / max_len
    cov_chain = max_len / cL if cL>0 else 0.0
    cov_parent = max_len / pL if pL>0 else 0.0
    return matches, max_len, identity, cov_chain, cov_parent, parent_start, chain_start, max_len

# ------------------------- 메인 -------------------------
def main():
    ap = argparse.ArgumentParser(description="parent full_fasta ↔ PDB 체인 매핑(k-mer/오프셋 투표, 실시간 append/flush)")
    ap.add_argument("--per-parent", required=True, help="per_parent_epitopes_clean.csv 또는 .xlsx 경로")
    ap.add_argument("--pdb-fasta", required=True, help="pdb_chains.fasta 경로 (>PDBID_CHAIN 헤더)")
    ap.add_argument("--out-best", default=None, help="최종 단일 매핑 CSV (기본: mapping_best.csv)")
    ap.add_argument("--out-all", default=None, help="임계치 통과 모든 후보 CSV (기본: mapping_all.csv)")
    ap.add_argument("--out-unmapped", default=None, help="매핑 실패 체인 CSV (기본: unmapped_chains.csv)")
    ap.add_argument("--append", action="store_true", help="기존 CSV에 이어쓰기(헤더 중복 방지)")
    ap.add_argument("--kmer", type=int, default=7, help="k-mer 길이 (기본 7)")
    ap.add_argument("--min-identity", type=float, default=0.80, help="최소 일치율 (기본 0.80)")
    ap.add_argument("--min-coverage", type=float, default=0.60, help="최소 커버리지(체인 기준) (기본 0.60)")
    ap.add_argument("--min-kmer-hits", type=int, default=3, help="최소 k-mer 히트 수 (기본 3)")
    ap.add_argument("--max-candidates", type=int, default=20, help="후보 평가 상한 (기본 20)")
    ap.add_argument("--exclude", default=None, help="제외할 parent 정규식 (예: (?i)\\bAra\\s*h\\s*2\\b)")
    ap.add_argument("--include", default=None, help="포함할 parent 정규식 (미지정 시 전체)")
    ap.add_argument("--min-len", type=int, default=30, help="너무 짧은 체인 제외 길이 하한 (기본 30)")
    ap.add_argument("--verbose", action="store_true", help="실행 과정 출력")
    ap.add_argument("--progress-every", type=int, default=1, help="진행 메시지 출력 주기(체인 단위)")
    ap.add_argument("--flush-every", type=int, default=1, help="CSV flush 주기(체인 단위, 기본 1)")
    ap.add_argument("--log-file", default=None, help="로그 파일 경로(콘솔+파일 동시 기록)")
    args = ap.parse_args()

    logger = TeeLogger(args.log_file)

    per_parent_path = Path(args.per_parent)
    pdb_fasta_path  = Path(args.pdb_fasta)

    out_best = Path(args.out_best) if args.out_best else pdb_fasta_path.with_name("mapping_best.csv")
    out_all  = Path(args.out_all)  if args.out_all  else pdb_fasta_path.with_name("mapping_all.csv")
    out_unm  = Path(args.out_unmapped) if args.out_unmapped else pdb_fasta_path.with_name("unmapped_chains.csv")

    # 1) per-parent 읽기
    parents = read_per_parent(per_parent_path)
    if args.include:
        pat_in = re.compile(args.include)
        parents = [p for p in parents if pat_in.search(p[0])]
    if args.exclude:
        pat_ex = re.compile(args.exclude)
        parents = [p for p in parents if not pat_ex.search(p[0])]
    if not parents:
        logger.log("ERROR: 사용할 parent 서열이 없습니다(파일 또는 필터를 확인).")
        logger.close(); sys.exit(2)

    # 2) PDB 체인 읽기
    if not pdb_fasta_path.exists():
        logger.log(f"ERROR: pdb 체인 FASTA가 없습니다: {pdb_fasta_path}")
        logger.log("팁) make_pdb_chains_from_pdb_seqres_ko.py로 생성하거나, 보유한 체인 FASTA 경로를 --pdb-fasta로 지정하세요.")
        logger.close(); sys.exit(2)
    chain_seqs = read_fasta(pdb_fasta_path)
    if not chain_seqs:
        logger.log("ERROR: pdb_chains.fasta에서 체인 서열을 읽지 못했습니다.")
        logger.close(); sys.exit(2)

    # 3) k-mer 인덱스
    if args.verbose:
        logger.log(f"[INFO] parents={len(parents)}  chains={len(chain_seqs)}  kmer={args.kmer}")
    index = build_kmer_index(parents, k=args.kmer)
    if args.verbose:
        logger.log(f"[INFO] k-mer index built. unique_kmers={len(index)}")

    # 4) CSV 준비 (append/overwrite)
    best_header = [
        "PDB chain","chain_len","parent_name","parent_len","offset",
        "kmer_hits","kmer_hits_at_best_offset","overlap_len",
        "identity","coverage_chain","coverage_parent",
        "parent_start","chain_start","method","pass"
    ]
    all_header = [
        "PDB chain","chain_len","parent_name","parent_len","offset",
        "kmer_hits","kmer_hits_at_best_offset","overlap_len",
        "identity","coverage_chain","coverage_parent",
        "parent_start","chain_start","method","pass"
    ]
    unm_header = ["PDB chain","chain_len","reason"]

    def open_csv_writer(path: Path, header: list, append: bool):
        mode = "a" if append and path.exists() else "w"
        f = open(path, mode, newline="", encoding="utf-8")
        w = csv.DictWriter(f, fieldnames=header)
        if mode == "w":
            w.writeheader(); f.flush()
        return f, w

    f_best, w_best = open_csv_writer(out_best, best_header, args.append)
    f_all,  w_all  = open_csv_writer(out_all,  all_header,  args.append)
    f_unm,  w_unm  = open_csv_writer(out_unm,  unm_header,  args.append)

    # 5) 체인별 매핑 (실시간 append/flush)
    headers = list(chain_seqs.keys())
    total = len(headers)
    flush_every = max(1, int(args.flush_every))
    processed = 0; mapped=0; candidates_total=0; unmapped=0

    for idx, header in enumerate(headers, 1):
        chain_id = header.strip()
        cseq = chain_seqs[header]
        cL = len(cseq)

        if cL < args.min_len:
            w_unm.writerow({"PDB chain": chain_id, "chain_len": cL, "reason": f"too_short(<{args.min_len})"})
            unmapped += 1
            processed += 1
            if args.verbose and (idx % args.progress_every == 0):
                logger.log(f"[{idx}/{total}] {chain_id} len={cL}  candidates=0")
                logger.log(f"  - unmapped: too_short(<{args.min_len})")
            if processed % flush_every == 0:
                f_unm.flush()
            continue

        # 후보 탐색
        votes, kmer_hits = vote_offsets_for_chain(cseq, index, args.kmer)
        candidates = sorted(kmer_hits.items(), key=lambda x: x[1], reverse=True)[:args.max_candidates]
        if args.verbose and (idx % args.progress_every == 0):
            logger.log(f"[{idx}/{total}] {chain_id} len={cL}  candidates={len(candidates)}")

        Scored = []
        for pi, hitcnt in candidates:
            pname, pseq = parents[pi]
            offs = votes[pi]
            if not offs:
                continue
            best_off, best_hits = offs.most_common(1)[0]
            matches, ovlen, ident, cov_chain, cov_parent, p_start, c_start, ovL = eval_alignment(pseq, cseq, best_off)

            row = {
                "PDB chain": chain_id,
                "chain_len": cL,
                "parent_name": pname,
                "parent_len": len(pseq),
                "offset": best_off,
                "kmer_hits": int(hitcnt),
                "kmer_hits_at_best_offset": int(best_hits),
                "overlap_len": int(ovlen),
                "identity": round(ident, 6),
                "coverage_chain": round(cov_chain, 6),
                "coverage_parent": round(cov_parent, 6),
                "parent_start": int(p_start),
                "chain_start": int(c_start),
                "method": "kmer_offset",
                "pass": bool((hitcnt >= args.min_kmer_hits) and (ident >= args.min_identity) and (cov_chain >= args.min_coverage))
            }
            Scored.append(row)
            if row["pass"]:
                w_all.writerow(row)

        # 즉시 flush(원하면 주기 조절 가능)
        candidates_total += sum(1 for r in Scored if r["pass"])
        if processed % flush_every == 0:
            f_all.flush()

        passed = [r for r in Scored if r["pass"]]
        if not passed:
            reason = "no_candidate" if not candidates else "below_threshold"
            w_unm.writerow({"PDB chain": chain_id, "chain_len": cL, "reason": reason})
            unmapped += 1
            if args.verbose and (idx % args.progress_every == 0):
                logger.log(f"  - unmapped: {reason}")
        else:
            best = sorted(passed, key=lambda r: (r["identity"], r["coverage_chain"], r["kmer_hits"]), reverse=True)[0]
            w_best.writerow(best)
            mapped += 1
            if args.verbose and (idx % args.progress_every == 0):
                logger.log(f"  - pick: parent='{best['parent_name']}'  ident={best['identity']:.3f}  cov_chain={best['coverage_chain']:.2f}  offset={best['offset']}  kmer_hits={best['kmer_hits']}")

        processed += 1
        # 파일 flush 주기
        if processed % flush_every == 0:
            f_best.flush(); f_unm.flush(); f_all.flush()

    # 마지막 flush
    f_best.flush(); f_unm.flush(); f_all.flush()
    f_best.close(); f_unm.close(); f_all.close()

    if args.verbose:
        logger.log(f"[DONE] chains={total}  mapped={mapped}  candidates={candidates_total}  unmapped={unmapped}")
        logger.log(f"       out_best={out_best}  out_all={out_all}  out_unmapped={out_unm}")
    logger.close()

if __name__ == "__main__":
    main()
